"""Export endpoints: Excel and PDF."""

import io
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.auth import get_current_user
from app.database import get_db
from app.models.financial import (
    Company, FinancialPeriod, FinancialData, ForecastScenario,
    LineItemDef, PeriodType, StatementType,
)
from app.models.user import User

router = APIRouter(
    prefix="/api/companies/{company_id}/scenarios/{scenario_id}/export",
    tags=["export"],
)


async def _load_statement_data(
    company_id: int, scenario_id: int, user: User, session: AsyncSession
):
    """Load all data needed for export: company, line items, period data."""
    company = await session.get(Company, company_id)
    if not company or company.owner_id != user.id:
        raise HTTPException(status_code=404, detail="Company not found")

    result = await session.execute(
        select(ForecastScenario).where(
            ForecastScenario.id == scenario_id,
            ForecastScenario.company_id == company_id,
        )
    )
    scenario = result.scalar_one_or_none()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")

    # Load periods
    result = await session.execute(
        select(FinancialPeriod)
        .where(
            FinancialPeriod.company_id == company_id,
            (
                (FinancialPeriod.period_type == PeriodType.HISTORICAL)
                | (FinancialPeriod.scenario_id == scenario_id)
            ),
        )
        .options(selectinload(FinancialPeriod.data))
        .order_by(FinancialPeriod.year)
    )
    periods = result.scalars().all()

    if not periods:
        raise HTTPException(status_code=404, detail="No data found")

    years = sorted({p.year for p in periods})
    period_types = {p.year: p.period_type.value for p in periods}
    data_map: dict[str, dict[int, float]] = {}
    for p in periods:
        for d in p.data:
            data_map.setdefault(d.line_item_code, {})[p.year] = d.value

    # Load line items
    lid_result = await session.execute(
        select(LineItemDef).order_by(LineItemDef.sort_order)
    )
    line_items = lid_result.scalars().all()

    return company, scenario, years, period_types, data_map, line_items


@router.get("/excel")
async def export_excel(
    company_id: int,
    scenario_id: int,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    company, scenario, years, period_types, data_map, line_items = (
        await _load_statement_data(company_id, scenario_id, current_user, session)
    )

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    hist_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    proj_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )
    bold_border = Border(
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )
    num_format = '#,##0'

    sheet_configs = [
        (StatementType.IS, "Income Statement"),
        (StatementType.BS, "Balance Sheet"),
        (StatementType.CF, "Cash Flow"),
    ]

    for st, sheet_name in sheet_configs:
        ws = wb.create_sheet(title=sheet_name)
        items = [li for li in line_items if li.statement_type == st]

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(years))
        title_cell = ws.cell(row=1, column=1, value=f"{company.name} — {sheet_name} ({scenario.name})")
        title_cell.font = header_font

        # Header row
        row = 3
        ws.cell(row=row, column=1, value="Line Item").font = header_font_white
        ws.cell(row=row, column=1).fill = header_fill
        ws.column_dimensions['A'].width = 35

        for ci, y in enumerate(years, start=2):
            label = f"{y} ({'Actual' if period_types.get(y) == 'HISTORICAL' else 'Forecast'})"
            cell = ws.cell(row=row, column=ci, value=label)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="right")
            ws.column_dimensions[cell.column_letter].width = 16

        # Data rows
        for li in items:
            row += 1
            # Name cell with indentation
            name = ("  " * li.indent_level) + li.name_cn
            name_cell = ws.cell(row=row, column=1, value=name)
            name_cell.font = bold_font if li.is_bold else normal_font
            name_cell.border = bold_border if li.is_bold else thin_border

            for ci, y in enumerate(years, start=2):
                val = data_map.get(li.code, {}).get(y, 0)
                cell = ws.cell(row=row, column=ci, value=val if val != 0 else None)
                cell.number_format = num_format
                cell.alignment = Alignment(horizontal="right")
                cell.font = bold_font if li.is_bold else normal_font
                cell.border = bold_border if li.is_bold else thin_border
                if period_types.get(y) == "HISTORICAL":
                    cell.fill = hist_fill
                else:
                    cell.fill = proj_fill

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{company.name}_{scenario.name}_forecast.xlsx"
    filename_encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"},
    )


@router.get("/pdf")
async def export_pdf(
    company_id: int,
    scenario_id: int,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    """Export PDF — generates a simple HTML-based PDF using basic table formatting."""
    company, scenario, years, period_types, data_map, line_items = (
        await _load_statement_data(company_id, scenario_id, current_user, session)
    )

    def fmt(v: float) -> str:
        if v == 0:
            return "-"
        s = f"{abs(v):,.0f}"
        return f"({s})" if v < 0 else s

    # Build HTML
    html_parts = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'>",
        "<style>",
        "body { font-family: 'Microsoft YaHei', 'SimSun', Arial, sans-serif; font-size: 10px; margin: 20px; }",
        "h1 { font-size: 16px; margin-bottom: 4px; }",
        "h2 { font-size: 13px; margin-top: 20px; color: #333; border-bottom: 1px solid #ccc; padding-bottom: 4px; }",
        "table { width: 100%; border-collapse: collapse; margin-bottom: 16px; }",
        "th { background: #1a73e8; color: #fff; padding: 6px 8px; text-align: right; font-size: 10px; }",
        "th:first-child { text-align: left; }",
        "td { padding: 4px 8px; border-bottom: 1px solid #eee; text-align: right; font-size: 10px; }",
        "td:first-child { text-align: left; }",
        ".bold { font-weight: bold; background: #f9f9f9; }",
        ".hist { background: #f5f5f5; }",
        ".proj { background: #e8f0fe; }",
        ".neg { color: #cf1322; }",
        ".cover { text-align: center; padding: 60px 0; }",
        ".cover h1 { font-size: 24px; }",
        ".cover p { color: #666; font-size: 12px; }",
        "@media print { .page-break { page-break-before: always; } }",
        "</style></head><body>",
    ]

    # Cover page
    html_parts.append(f"""
        <div class="cover">
            <h1>{company.name}</h1>
            <p>Financial Forecast Report</p>
            <p>Scenario: {scenario.name}</p>
            <p>Years: {', '.join(str(y) for y in years)}</p>
        </div>
    """)

    sheet_configs = [
        (StatementType.IS, "Income Statement / 利润表"),
        (StatementType.BS, "Balance Sheet / 资产负债表"),
        (StatementType.CF, "Cash Flow / 现金流量表"),
    ]

    for st, title in sheet_configs:
        items = [li for li in line_items if li.statement_type == st]
        html_parts.append('<div class="page-break"></div>')
        html_parts.append(f"<h2>{title}</h2>")
        html_parts.append("<table><thead><tr><th>Line Item</th>")
        for y in years:
            pt = "Actual" if period_types.get(y) == "HISTORICAL" else "Forecast"
            html_parts.append(f"<th>{y}<br/><small>{pt}</small></th>")
        html_parts.append("</tr></thead><tbody>")

        for li in items:
            cls = ' class="bold"' if li.is_bold else ""
            html_parts.append(f"<tr{cls}>")
            indent = "&nbsp;&nbsp;" * li.indent_level
            html_parts.append(f"<td>{indent}{li.name_cn}</td>")
            for y in years:
                val = data_map.get(li.code, {}).get(y, 0)
                pt_cls = "hist" if period_types.get(y) == "HISTORICAL" else "proj"
                neg_cls = " neg" if val < 0 else ""
                html_parts.append(f'<td class="{pt_cls}{neg_cls}">{fmt(val)}</td>')
            html_parts.append("</tr>")

        html_parts.append("</tbody></table>")

    html_parts.append("</body></html>")
    html_content = "\n".join(html_parts)

    buf = io.BytesIO(html_content.encode("utf-8"))
    filename = f"{company.name}_{scenario.name}_forecast.html"
    filename_encoded = quote(filename)

    return StreamingResponse(
        buf,
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"},
    )
